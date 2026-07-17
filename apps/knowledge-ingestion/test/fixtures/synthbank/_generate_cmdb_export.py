#!/usr/bin/env python3
"""Generate synthbank-cmdb-export.xlsx — the ~260-CI SynthBank estate (P1 CMDB sibling).

SYNTHETIC: SynthBank is a fictional medium UCB. Every name, owner, ref and figure
is fabricated. This is the CMDB-export sibling to the Stage 1 doc set; it shares
tenant cfc5801f-... and the Stage 1 CI vocabulary (superset of the 16-CI W2 fixture).

Ingest reality (Phase 0, confirmed): the W2 path (xlsx.parser) reads ONLY the first
worksheet and the chunker emits 1 chunk per row. So:
  * Sheet 1 'configuration_items'  -> INGESTED  (~256 CI rows -> ~256 chunks)
  * Sheets 2-5 (services/links/relationships/change_links) -> NOT INGESTED by W2.
    They carry the full relational spine in the file, ready for the deferred
    DataSourceProvider-mediated import (Phase 2 / W6, flag-don't-build per Phase 0).

Run:  python3 _generate_cmdb_export.py
"""
import os
from openpyxl import Workbook

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "synthbank-cmdb-export.xlsx")

# Sheet-1 columns: superset of the 16-CI fixture (ci_id..location) + brief columns.
CI_COLUMNS = [
    "ci_id", "ci_name", "ci_type", "criticality_tier", "business_service",
    "location", "linked_asset_ref", "technical_owner", "business_owner",
    "operations_team", "status", "dr_mapping",
]

# ---------------------------------------------------------------------------
# 6 regional hubs and the contiguous branch -> hub assignment.
# ---------------------------------------------------------------------------
HUBS = ["Hub-North", "Hub-South", "Hub-East", "Hub-West", "Hub-Central", "Hub-NorthEast"]
# branch i (1..50) -> hub by contiguous block (9,9,9,9,9,5). Branch-023 -> Hub-East.
def hub_for_branch(i):
    if i <= 9:   return "Hub-North"
    if i <= 18:  return "Hub-South"
    if i <= 27:  return "Hub-East"
    if i <= 36:  return "Hub-West"
    if i <= 45:  return "Hub-Central"
    return "Hub-NorthEast"

rows = []          # list of dicts keyed by CI_COLUMNS
imperfections = [] # audit log lines
_next = [17]       # ci_id counter; CI-0001..0016 are the legacy overlap, new start 0017

def cid_new():
    n = _next[0]; _next[0] += 1
    return f"CI-{n:04d}"

def add(ci_id, name, ci_type, tier, service, location, asset_ref,
        tech_owner, biz_owner, ops_team, status="active", dr=""):
    rows.append({
        "ci_id": ci_id, "ci_name": name, "ci_type": ci_type,
        "criticality_tier": tier, "business_service": service, "location": location,
        "linked_asset_ref": asset_ref, "technical_owner": tech_owner,
        "business_owner": biz_owner, "operations_team": ops_team,
        "status": status, "dr_mapping": dr,
    })

# ---------------------------------------------------------------------------
# (A) DC-Primary + DC-DR + WAN-edge singletons.
#     The first 16 reproduce the legacy 16-CI fixture verbatim (same id/name/
#     type/tier/service/location) so the estate is a clean superset; the new
#     columns are added. Overlap is by design (manifest-noted, additive).
# ---------------------------------------------------------------------------
# dr_mapping wired here; atm_card_services CIs deliberately left blank (the gap).
add("CI-0001", "CBS App Node 1", "cbs_application_server", "tier_1", "core_banking", "DC-Primary",
    "ems-asset://dc1/cbs-app-01", "core.ops@synthbank.example", "head.cbs@synthbank.example", "core-banking-ops",
    dr="CBS App DR Node 1")
add("CI-0002", "CBS DB Node 1", "cbs_database_server", "tier_1", "core_banking", "DC-Primary",
    "ems-asset://dc1/cbs-db-01", "dba.team@synthbank.example", "head.cbs@synthbank.example", "core-banking-ops",
    dr="CBS DB DR Node 1")
add("CI-0003", "CBS Hosted Service", "cbs_hosted_service", "tier_1", "core_banking", "Hosted-Shared",
    "", "vendor.mgmt@synthbank.example", "head.cbs@synthbank.example", "vendor-management",
    dr="")  # hosted/shared — DR handled by vendor contract (intentionally blank ref)
add("CI-0004", "UPI Switch 1", "upi_switch", "tier_1", "upi_imps", "DC-Primary",
    "ems-asset://dc1/upi-sw-01", "payments.ops@synthbank.example", "head.payments@synthbank.example", "payments-ops",
    dr="UPI Switch 2 (DR)")
add("CI-0005", "Sponsor Bank Link A", "sponsor_bank_link", "tier_1", "upi_imps", "WAN-Edge",
    "ems-asset://edge/sponsor-link-a", "payments.ops@synthbank.example", "head.payments@synthbank.example", "payments-ops",
    dr="Sponsor Bank Link B (DR)")
add("CI-0006", "NPCI Link A", "npci_link", "tier_1", "upi_imps", "WAN-Edge",
    "ems-asset://edge/npci-link-a", "payments.ops@synthbank.example", "head.payments@synthbank.example", "payments-ops",
    dr="NPCI Link B (Secondary)")
add("CI-0007", "Payment Gateway 1", "payment_gateway", "tier_1", "upi_imps", "DC-Primary",
    "ems-asset://dc1/pg-01", "payments.ops@synthbank.example", "head.payments@synthbank.example", "payments-ops",
    dr="Payment Gateway 2 (DR)")
# --- atm_card_services: NO DR (imperfection #5 — aligns with Stage 1 doc gap) ---
add("CI-0008", "ATM Switch 1", "atm_switch", "tier_1", "atm_card_services", "DC-Primary",
    "ems-asset://dc1/atm-sw-01", "channels.ops@synthbank.example", "head.cards@synthbank.example", "channels-ops",
    dr="")  # deliberate: no DR node for atm_card_services
add("CI-0009", "HSM Device 1", "hsm_device", "tier_1", "atm_card_services", "DC-Primary",
    "ems-asset://dc1/hsm-01", "security.ops@synthbank.example", "head.cards@synthbank.example", "security-ops",
    dr="")  # deliberate: no DR for the card-serving HSM
add("CI-0010", "Sponsor Bank Link B (DR)", "sponsor_bank_link", "tier_1", "neft_rtgs", "DC-DR",
    "ems-asset://edge/sponsor-link-b", "payments.ops@synthbank.example", "head.payments@synthbank.example", "payments-ops",
    dr="")  # this IS the DR node
add("CI-0011", "Core Router 1", "core_router", "tier_2", "branch_wan", "DC-Primary",
    "ems-asset://dc1/core-rtr-01", "network.ops@synthbank.example", "head.branch.banking@synthbank.example", "network-ops",
    dr="Core Router DR")
add("CI-0012", "Branch Router BR-014", "branch_router", "tier_2", "branch_wan", "Branch-014",
    "ems-asset://branch-014/rtr", "network.ops@synthbank.example", "regional.head.south@synthbank.example", "network-ops",
    dr="")  # branch CIs have no DR node
add("CI-0013", "Firewall Edge 1", "firewall", "tier_2", "branch_wan", "WAN-Edge",
    "ems-asset://edge/fw-01", "security.ops@synthbank.example", "head.branch.banking@synthbank.example", "security-ops",
    dr="Firewall DR")
add("CI-0014", "DR Site Node 1", "dr_site_node", "tier_1", "core_banking", "DC-DR",
    "ems-asset://dc-dr/dr-node-01", "bcp.team@synthbank.example", "head.cbs@synthbank.example", "bcp-team",
    dr="")  # this IS a DR node
add("CI-0015", "Backup System 1", "backup_system", "tier_3", "document_management", "DC-Primary",
    "ems-asset://dc1/backup-01", "infra.ops@synthbank.example", "cio.office@synthbank.example", "infra-ops",
    dr="Backup System 2 (DR)")
add("CI-0016", "Regulatory Reporting Server", "cbs_application_server", "tier_2", "regulatory_reporting", "DC-Primary",
    "ems-asset://dc1/regrep-01", "compliance.ops@synthbank.example", "head.compliance@synthbank.example", "compliance-ops",
    dr="")  # tier-2; no DR modelled

# --- New DC-Primary CIs (CI-0017+) extending the singletons to a believable DC ---
add(cid_new(), "CBS App Node 2", "cbs_application_server", "tier_1", "core_banking", "DC-Primary",
    "ems-asset://dc1/cbs-app-02", "core.ops@synthbank.example", "head.cbs@synthbank.example", "core-banking-ops",
    dr="CBS App DR Node 1")
add(cid_new(), "CBS DB Node 2", "cbs_database_server", "tier_1", "core_banking", "DC-Primary",
    "ems-asset://dc1/cbs-db-02", "dba.team@synthbank.example", "head.cbs@synthbank.example", "core-banking-ops",
    dr="CBS DB DR Node 1")
add(cid_new(), "Core Switch 1", "core_switch", "tier_2", "core_banking", "DC-Primary",
    "ems-asset://dc1/core-sw-01", "network.ops@synthbank.example", "head.cbs@synthbank.example", "network-ops",
    dr="Core Switch DR")
add(cid_new(), "Core Switch 2", "core_switch", "tier_2", "core_banking", "DC-Primary",
    "ems-asset://dc1/core-sw-02", "network.ops@synthbank.example", "head.cbs@synthbank.example", "network-ops",
    dr="Core Switch DR")
add(cid_new(), "Core Router 2", "core_router", "tier_2", "branch_wan", "DC-Primary",
    "ems-asset://dc1/core-rtr-02", "network.ops@synthbank.example", "head.branch.banking@synthbank.example", "network-ops",
    dr="Core Router DR")
add(cid_new(), "Firewall Edge 2", "firewall", "tier_2", "internet_mobile_banking", "WAN-Edge",
    "ems-asset://edge/fw-02", "security.ops@synthbank.example", "head.channels@synthbank.example", "security-ops",
    dr="Firewall DR")
add(cid_new(), "HSM Device 2", "hsm_device", "tier_1", "upi_imps", "DC-Primary",
    "ems-asset://dc1/hsm-02", "security.ops@synthbank.example", "head.payments@synthbank.example", "security-ops",
    dr="")  # HSM-2 serves UPI; UPI DR is via switch/PG, so blank here is fine
add(cid_new(), "AD/DNS/DHCP Server 1", "ad_dns_dhcp", "tier_2", "core_banking", "DC-Primary",
    "ems-asset://dc1/addns-01", "infra.ops@synthbank.example", "cio.office@synthbank.example", "infra-ops",
    dr="AD/DNS/DHCP Server 2 (DR)")
add(cid_new(), "Internet Banking Server 1", "internet_banking_server", "tier_1", "internet_mobile_banking", "DC-Primary",
    "ems-asset://dc1/ibank-01", "channels.ops@synthbank.example", "head.channels@synthbank.example", "channels-ops",
    dr="Internet Banking Server 2 (DR)")
add(cid_new(), "Mobile Banking Gateway 1", "mobile_banking_gateway", "tier_1", "internet_mobile_banking", "DC-Primary",
    "ems-asset://dc1/mbank-01", "channels.ops@synthbank.example", "head.channels@synthbank.example", "channels-ops",
    dr="Mobile Banking Gateway 2 (DR)")
add(cid_new(), "CTS System 1", "cts_system", "tier_1", "cheque_clearing", "DC-Primary",
    "ems-asset://dc1/cts-01", "core.ops@synthbank.example", "head.cbs@synthbank.example", "core-banking-ops",
    dr="CTS System 2 (DR)")
add(cid_new(), "Backup System 2 (DR)", "backup_system", "tier_3", "document_management", "DC-DR",
    "ems-asset://dc-dr/backup-02", "infra.ops@synthbank.example", "cio.office@synthbank.example", "infra-ops",
    dr="")

# --- DC-DR mirror of critical CIs (no atm_switch / card-HSM DR — the gap) ---
add(cid_new(), "CBS App DR Node 1", "cbs_application_server", "tier_1", "core_banking", "DC-DR",
    "ems-asset://dc-dr/cbs-app-01", "bcp.team@synthbank.example", "head.cbs@synthbank.example", "bcp-team")
add(cid_new(), "CBS DB DR Node 1", "cbs_database_server", "tier_1", "core_banking", "DC-DR",
    "ems-asset://dc-dr/cbs-db-01", "bcp.team@synthbank.example", "head.cbs@synthbank.example", "bcp-team")
add(cid_new(), "UPI Switch 2 (DR)", "upi_switch", "tier_1", "upi_imps", "DC-DR",
    "ems-asset://dc-dr/upi-sw-02", "payments.ops@synthbank.example", "head.payments@synthbank.example", "payments-ops")
add(cid_new(), "Payment Gateway 2 (DR)", "payment_gateway", "tier_1", "upi_imps", "DC-DR",
    "ems-asset://dc-dr/pg-02", "payments.ops@synthbank.example", "head.payments@synthbank.example", "payments-ops")
add(cid_new(), "Internet Banking Server 2 (DR)", "internet_banking_server", "tier_1", "internet_mobile_banking", "DC-DR",
    "ems-asset://dc-dr/ibank-02", "channels.ops@synthbank.example", "head.channels@synthbank.example", "channels-ops")
add(cid_new(), "Mobile Banking Gateway 2 (DR)", "mobile_banking_gateway", "tier_1", "internet_mobile_banking", "DC-DR",
    "ems-asset://dc-dr/mbank-02", "channels.ops@synthbank.example", "head.channels@synthbank.example", "channels-ops")
add(cid_new(), "CTS System 2 (DR)", "cts_system", "tier_1", "cheque_clearing", "DC-DR",
    "ems-asset://dc-dr/cts-02", "core.ops@synthbank.example", "head.cbs@synthbank.example", "core-banking-ops")
add(cid_new(), "Core Router DR", "core_router", "tier_2", "branch_wan", "DC-DR",
    "ems-asset://dc-dr/core-rtr", "network.ops@synthbank.example", "head.branch.banking@synthbank.example", "network-ops")
add(cid_new(), "Core Switch DR", "core_switch", "tier_2", "core_banking", "DC-DR",
    "ems-asset://dc-dr/core-sw", "network.ops@synthbank.example", "head.cbs@synthbank.example", "network-ops")
add(cid_new(), "Firewall DR", "firewall", "tier_2", "branch_wan", "DC-DR",
    "ems-asset://dc-dr/fw", "security.ops@synthbank.example", "head.branch.banking@synthbank.example", "security-ops")
add(cid_new(), "AD/DNS/DHCP Server 2 (DR)", "ad_dns_dhcp", "tier_2", "core_banking", "DC-DR",
    "ems-asset://dc-dr/addns-02", "infra.ops@synthbank.example", "cio.office@synthbank.example", "infra-ops")

# --- The deliberate unmonitored secondary rail (imperfection #2: partial) ---
add(cid_new(), "NPCI Link B (Secondary)", "npci_link", "tier_2", "upi_imps", "WAN-Edge",
    "", "payments.ops@synthbank.example", "head.payments@synthbank.example", "payments-ops",
    dr="")  # blank linked_asset_ref -> unmonitored secondary rail (no metrics)
imperfections.append("Unmonitored secondary rail: 'NPCI Link B (Secondary)' has blank linked_asset_ref (no metrics) -> cmdb_context.completeness = partial.")

# ---------------------------------------------------------------------------
# (B) 6 regional hubs (hub_router + hub_switch each).
# ---------------------------------------------------------------------------
for h, hub in enumerate(HUBS, start=1):
    add(cid_new(), f"{hub} Router", "hub_router", "tier_2", "branch_wan", hub,
        f"ems-asset://{hub.lower()}/rtr", "network.ops@synthbank.example",
        "head.branch.banking@synthbank.example", "network-ops")
    add(cid_new(), f"{hub} Switch", "hub_switch", "tier_2", "branch_wan", hub,
        f"ems-asset://{hub.lower()}/sw", "network.ops@synthbank.example",
        "head.branch.banking@synthbank.example", "network-ops")

# ---------------------------------------------------------------------------
# (C) 50 branches x (branch_router, branch_switch, 2x atm_terminal) = 200 CIs.
#     Branch-014's router reuses the legacy CI-0012 above, so skip it here.
#     2 branch switches get criticality_tier='unknown' (imperfection #4).
# ---------------------------------------------------------------------------
UNKNOWN_TIER_BRANCHES = {7, 31}  # imperfection #4: 2 CIs with unknown tier
for i in range(1, 51):
    loc = f"Branch-{i:03d}"
    hub = hub_for_branch(i)
    reg = {"Hub-North": "north", "Hub-South": "south", "Hub-East": "east",
           "Hub-West": "west", "Hub-Central": "central", "Hub-NorthEast": "northeast"}[hub]
    biz_owner = f"regional.head.{reg}@synthbank.example"
    # branch_router (skip 14 — already added as legacy CI-0012)
    if i != 14:
        add(cid_new(), f"Branch Router BR-{i:03d}", "branch_router", "tier_2", "branch_wan", loc,
            f"ems-asset://branch-{i:03d}/rtr", "network.ops@synthbank.example", biz_owner, "network-ops")
    # branch_switch
    sw_tier = "unknown" if i in UNKNOWN_TIER_BRANCHES else "tier_3"
    add(cid_new(), f"Branch Switch SW-{i:03d}", "branch_switch", sw_tier, "branch_wan", loc,
        f"ems-asset://branch-{i:03d}/sw", "network.ops@synthbank.example", biz_owner, "network-ops")
    # 2 atm_terminals (atm_card_services -> no DR by design)
    for t in (1, 2):
        add(cid_new(), f"ATM Terminal ATM-{i:03d}-{t}", "atm_terminal", "tier_2", "atm_card_services", loc,
            f"ems-asset://branch-{i:03d}/atm-{t}", "channels.ops@synthbank.example", biz_owner, "channels-ops")

for b in sorted(UNKNOWN_TIER_BRANCHES):
    imperfections.append(f"Unknown criticality_tier: 'Branch Switch SW-{b:03d}' left tier=unknown -> tier-unknown handling.")

# ---------------------------------------------------------------------------
# (D) Orphaned CIs — blank business_service (imperfection #1, 4 orphans incl.
#     the Stage 1 'Reconciliation Server 1' reference).
# ---------------------------------------------------------------------------
add(cid_new(), "Reconciliation Server 1", "recon_server", "tier_2", "", "DC-Primary",
    "ems-asset://dc1/recon-01", "payments.ops@synthbank.example", "", "payments-ops")
add(cid_new(), "Legacy Reporting Server", "cbs_application_server", "tier_3", "", "DC-Primary",
    "", "infra.ops@synthbank.example", "", "infra-ops", status="retired")
add(cid_new(), "Lab Sandbox Server", "server", "unknown", "", "DC-Primary",
    "", "infra.ops@synthbank.example", "", "infra-ops")
add(cid_new(), "Spare Edge Switch", "core_switch", "tier_3", "", "WAN-Edge",
    "ems-asset://edge/spare-sw", "network.ops@synthbank.example", "", "network-ops")
imperfections.append("Orphaned CIs (blank business_service -> 'unknown upstream' hygiene): "
                     "'Reconciliation Server 1' (the Stage 1 recon-SOP reference), 'Legacy Reporting Server' (retired), "
                     "'Lab Sandbox Server', 'Spare Edge Switch'. 4 total.")
imperfections.append("DR gap: atm_card_services (ATM Switch 1, HSM Device 1, all ATM terminals) has blank dr_mapping and NO "
                     "atm_switch/card-HSM node in DC-DR -> BCP-hygiene flag. Aligned with Stage 1 doc gap (imperfection #1: "
                     "atm_card_services DR procedure deliberately absent across the doc corpus).")

# ---------------------------------------------------------------------------
# (E) ~15% ownership incompleteness — deterministically blank one owner field
#     on every 7th CI (imperfection #3). Skip rows already intentionally blank.
# ---------------------------------------------------------------------------
owner_blanked = 0
for idx, r in enumerate(rows):
    if r["business_service"] == "":  # orphans already have blank biz_owner; don't double-count
        continue
    if idx % 7 == 3:
        if r["technical_owner"]:
            r["technical_owner"] = ""
            owner_blanked += 1
already_missing = sum(1 for r in rows if not r["technical_owner"] or not r["business_owner"])
imperfections.append(f"Ownership incompleteness: {already_missing} CIs ({already_missing*100//len(rows)}%) missing "
                     f"technical_owner or business_owner ({owner_blanked} technical_owner fields blanked deterministically "
                     f"+ orphan/hosted rows) -> ownership-completeness flag.")

# ---------------------------------------------------------------------------
# SHEET 2-5: the relational spine (NOT ingested by W2; ready for Phase 2/W6).
# ---------------------------------------------------------------------------
SERVICE_COLUMNS = ["service_name", "criticality_tier", "rto_minutes", "rpo_minutes",
                   "revenue_impact_hourly_inr", "business_owner", "description"]
SERVICES = [
    ["core_banking", "tier_1", 60, 15, 2500000, "head.cbs@synthbank.example", "CBS — accounts, GL, transaction processing [verify]"],
    ["upi_imps", "tier_1", 30, 5, 1800000, "head.payments@synthbank.example", "Retail digital payments via sponsor/NPCI [verify]"],
    ["neft_rtgs", "tier_1", 60, 15, 1200000, "head.payments@synthbank.example", "Interbank transfers (sub-membership) [verify]"],
    ["atm_card_services", "tier_1", 0, 0, 900000, "head.cards@synthbank.example", "ATM switching + card auth. NO DR documented (deliberate gap) [verify]"],
    ["internet_mobile_banking", "tier_1", 45, 15, 1100000, "head.channels@synthbank.example", "Internet + mobile banking channels [verify]"],
    ["cheque_clearing", "tier_1", 120, 60, 400000, "head.cbs@synthbank.example", "CTS cheque clearing [verify]"],
    ["branch_wan", "tier_2", 120, 60, 600000, "head.branch.banking@synthbank.example", "MPLS/VPN connectivity to branches [verify]"],
    ["regulatory_reporting", "tier_2", 240, 120, 150000, "head.compliance@synthbank.example", "RBI returns / ADF [verify]"],
    ["document_management", "tier_3", 480, 240, 0, "cio.office@synthbank.example", "Back-office document store / backup [verify]"],
]
imperfections.append("RTO/RPO=0 row for atm_card_services in the services sheet encodes the same DR gap relationally "
                     "(BCP posture violation flag).")

# service -> CI links (M:N). role in primary/backup/dependency.
LINK_COLUMNS = ["service_name", "ci_name", "role"]
LINKS = [
    # core_banking
    ("core_banking", "CBS App Node 1", "primary"), ("core_banking", "CBS App Node 2", "primary"),
    ("core_banking", "CBS DB Node 1", "primary"),  ("core_banking", "CBS DB Node 2", "primary"),
    ("core_banking", "CBS Hosted Service", "dependency"),
    ("core_banking", "Core Switch 1", "dependency"), ("core_banking", "Core Switch 2", "dependency"),
    ("core_banking", "CBS App DR Node 1", "backup"), ("core_banking", "CBS DB DR Node 1", "backup"),
    ("core_banking", "DR Site Node 1", "backup"),
    # upi_imps
    ("upi_imps", "UPI Switch 1", "primary"), ("upi_imps", "Payment Gateway 1", "primary"),
    ("upi_imps", "Sponsor Bank Link A", "dependency"), ("upi_imps", "NPCI Link A", "dependency"),
    ("upi_imps", "NPCI Link B (Secondary)", "dependency"), ("upi_imps", "HSM Device 2", "dependency"),
    ("upi_imps", "CBS App Node 1", "dependency"),
    ("upi_imps", "UPI Switch 2 (DR)", "backup"), ("upi_imps", "Payment Gateway 2 (DR)", "backup"),
    # neft_rtgs
    ("neft_rtgs", "Sponsor Bank Link A", "dependency"), ("neft_rtgs", "Sponsor Bank Link B (DR)", "backup"),
    ("neft_rtgs", "NPCI Link A", "dependency"), ("neft_rtgs", "CBS App Node 1", "dependency"),
    # atm_card_services (NO backup role — the deliberate DR gap)
    ("atm_card_services", "ATM Switch 1", "primary"), ("atm_card_services", "HSM Device 1", "dependency"),
    ("atm_card_services", "Sponsor Bank Link A", "dependency"),
    # internet_mobile_banking
    ("internet_mobile_banking", "Internet Banking Server 1", "primary"),
    ("internet_mobile_banking", "Mobile Banking Gateway 1", "primary"),
    ("internet_mobile_banking", "Firewall Edge 2", "dependency"), ("internet_mobile_banking", "CBS App Node 1", "dependency"),
    ("internet_mobile_banking", "Internet Banking Server 2 (DR)", "backup"),
    ("internet_mobile_banking", "Mobile Banking Gateway 2 (DR)", "backup"),
    # cheque_clearing
    ("cheque_clearing", "CTS System 1", "primary"), ("cheque_clearing", "Sponsor Bank Link A", "dependency"),
    ("cheque_clearing", "CTS System 2 (DR)", "backup"),
    # branch_wan
    ("branch_wan", "Core Router 1", "primary"), ("branch_wan", "Core Router 2", "primary"),
    ("branch_wan", "Firewall Edge 1", "dependency"), ("branch_wan", "Core Router DR", "backup"),
    # regulatory_reporting / document_management
    ("regulatory_reporting", "Regulatory Reporting Server", "primary"),
    ("document_management", "Backup System 1", "primary"), ("document_management", "Backup System 2 (DR)", "backup"),
]
# hub routers/switches and branch routers all serve branch_wan (dependency)
for r in rows:
    if r["ci_type"] in ("hub_router", "hub_switch", "branch_router"):
        LINKS.append(("branch_wan", r["ci_name"], "dependency"))
    if r["ci_type"] == "atm_terminal":
        LINKS.append(("atm_card_services", r["ci_name"], "dependency"))

# CI -> CI relationships (semantic). type in runs_on/depends_on/connected_to/hosts/contains.
REL_COLUMNS = ["source_ci", "target_ci", "relationship_type"]
RELATIONSHIPS = [
    ("CBS App Node 1", "CBS DB Node 1", "depends_on"),
    ("CBS App Node 2", "CBS DB Node 2", "depends_on"),
    ("UPI Switch 1", "Sponsor Bank Link A", "depends_on"),
    ("UPI Switch 1", "NPCI Link A", "depends_on"),
    ("UPI Switch 1", "HSM Device 2", "depends_on"),
    ("Payment Gateway 1", "UPI Switch 1", "depends_on"),
    ("ATM Switch 1", "HSM Device 1", "depends_on"),
    ("ATM Switch 1", "Sponsor Bank Link A", "depends_on"),
    ("Internet Banking Server 1", "CBS App Node 1", "depends_on"),
    ("Mobile Banking Gateway 1", "CBS App Node 1", "depends_on"),
    ("CTS System 1", "Sponsor Bank Link A", "depends_on"),
    ("Core Router 1", "Firewall Edge 1", "connected_to"),
]
# access-layer topology edges: atm->switch, switch->router, router->hub, hub->core
for i in range(1, 51):
    rtr = "Branch Router BR-014" if i == 14 else f"Branch Router BR-{i:03d}"
    sw = f"Branch Switch SW-{i:03d}"
    hub = hub_for_branch(i)
    RELATIONSHIPS.append((sw, rtr, "connected_to"))
    RELATIONSHIPS.append((rtr, f"{hub} Router", "connected_to"))
    for t in (1, 2):
        RELATIONSHIPS.append((f"ATM Terminal ATM-{i:03d}-{t}", sw, "connected_to"))
for hub in HUBS:
    RELATIONSHIPS.append((f"{hub} Router", "Core Router 1", "connected_to"))

# change records -> CIs (sheet 5). One is the RCA-2 smoking gun.
CHG_COLUMNS = ["change_ref", "ci_name", "change_role", "change_date", "risk", "summary"]
CHANGES = [
    ("CHG-SB-2026-0291", "CBS DB Node 1", "modified", "2026-03-14 21:30", "low",
     "Reduced CBS batch parallel-worker/connection-pool parameter (tuning). Assessed low-risk. "
     "SMOKING GUN for INC-SB-2026-0388 (RCA-2 EOD overrun the following early morning) [verify]."),
    ("CHG-SB-2026-0288", "UPI Switch 1", "modified", "2026-03-10 02:00", "medium",
     "UPI switch firmware patch during maintenance window [verify]."),
    ("CHG-SB-2026-0305", "Firewall Edge 1", "modified", "2026-03-20 23:15", "medium",
     "Edge firewall ruleset update for branch WAN segment [verify]."),
]
imperfections.append("Change record CHG-SB-2026-0291 (CBS DB Node 1, 2026-03-14 21:30, low-risk) is the deliberate "
                     "'recent change is the smoking gun' fuel for RCA-2/INC-SB-2026-0388 (W8). Lives in the change_links "
                     "sheet (NOT ingested by W2 — Phase 2/W6 relational import); the RCA-2 doc prose already carries the "
                     "retrievable narrative.")


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "configuration_items"
    ws.append(CI_COLUMNS)
    for r in rows:
        ws.append([r[c] for c in CI_COLUMNS])

    ws2 = wb.create_sheet("business_services")
    ws2.append(SERVICE_COLUMNS)
    for s in SERVICES:
        ws2.append(s)

    ws3 = wb.create_sheet("service_ci_links")
    ws3.append(LINK_COLUMNS)
    for l in LINKS:
        ws3.append(list(l))

    ws4 = wb.create_sheet("relationships")
    ws4.append(REL_COLUMNS)
    for rel in RELATIONSHIPS:
        ws4.append(list(rel))

    ws5 = wb.create_sheet("change_links")
    ws5.append(CHG_COLUMNS)
    for c in CHANGES:
        ws5.append(list(c))

    wb.save(OUT)

    # ---- audit summary ----
    by_loc = {}
    by_type = {}
    for r in rows:
        loc = "Branch" if r["location"].startswith("Branch-") else (
              "Hub" if r["location"].startswith("Hub-") else r["location"])
        by_loc[loc] = by_loc.get(loc, 0) + 1
        by_type[r["ci_type"]] = by_type.get(r["ci_type"], 0) + 1
    print(f"wrote {OUT}")
    print(f"SHEET 1 'configuration_items': {len(rows)} CI rows -> ~{len(rows)} chunks (1/row)")
    print(f"  by location group: {dict(sorted(by_loc.items()))}")
    print(f"  ci_type count: {len(by_type)} distinct")
    print(f"SHEET 2 'business_services': {len(SERVICES)} services")
    print(f"SHEET 3 'service_ci_links': {len(LINKS)} links")
    print(f"SHEET 4 'relationships': {len(RELATIONSHIPS)} edges")
    print(f"SHEET 5 'change_links': {len(CHANGES)} change records")
    print("\nIMPERFECTION LOG:")
    for line in imperfections:
        print("  - " + line)


if __name__ == "__main__":
    main()
